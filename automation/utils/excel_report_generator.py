import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from automation.config.config import EXCEL_RESULTS_DIR, REPORTS_DIR
from automation.utils.logger import get_logger

logger = get_logger("ExcelReportGenerator")

# Styling constants
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

PASS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
PASS_FONT = Font(name="Calibri", size=10, color="166534", bold=True)

FAIL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
FAIL_FONT = Font(name="Calibri", size=10, color="991B1B", bold=True)

SKIP_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
SKIP_FONT = Font(name="Calibri", size=10, color="854D0E", bold=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

class ExcelReportGenerator:
    def __init__(self, test_results: list[dict], summary_metrics: dict):
        self.results = test_results
        self.metrics = summary_metrics

    def _auto_adjust_columns(self, sheet):
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                max_len = max(max_len, len(val))
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 14)

    def _apply_row_style(self, row, fill=None, font=None, alignment=None):
        for cell in row:
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment

    def generate_all_reports(self) -> dict[str, Path]:
        """Generates all 4 required Excel reports."""
        generated_files = {}
        
        main_report = self.generate_main_automation_report()
        failed_report = self.generate_filtered_report("Failed", [r for r in self.results if r["status"] == "FAIL"], "Failed_Test_Cases.xlsx")
        passed_report = self.generate_filtered_report("Passed", [r for r in self.results if r["status"] == "PASS"], "Passed_Test_Cases.xlsx")
        summary_report = self.generate_summary_report("Summary_Report.xlsx")

        generated_files["main"] = main_report
        generated_files["failed"] = failed_report
        generated_files["passed"] = passed_report
        generated_files["summary"] = summary_report
        
        return generated_files

    def generate_main_automation_report(self) -> Path:
        """Generates Automation_Test_Report.xlsx with 6 comprehensive sheets."""
        wb = openpyxl.Workbook()
        
        # Sheet 1: Executed Test Cases
        ws1 = wb.active
        ws1.title = "Executed Test Cases"
        headers1 = ["Test ID", "Module", "Test Name", "Priority", "Status", "Execution Time (s)", "Timestamp"]
        ws1.append(headers1)
        self._apply_row_style(ws1[1], fill=HEADER_FILL, font=HEADER_FONT, alignment=Alignment(horizontal="center"))
        
        for r in self.results:
            row_data = [r["test_id"], r["module"], r["name"], r["priority"], r["status"], r.get("duration", 0.0), r.get("timestamp", "")]
            ws1.append(row_data)
            row_idx = ws1.max_row
            status_cell = ws1.cell(row=row_idx, column=5)
            self._apply_row_style(ws1[row_idx], alignment=Alignment(horizontal="left"))
            
            if r["status"] == "PASS":
                status_cell.fill = PASS_FILL
                status_cell.font = PASS_FONT
            elif r["status"] == "FAIL":
                status_cell.fill = FAIL_FILL
                status_cell.font = FAIL_FONT
            else:
                status_cell.fill = SKIP_FILL
                status_cell.font = SKIP_FONT
        self._auto_adjust_columns(ws1)

        # Sheet 2: Passed Tests
        ws2 = wb.create_sheet(title="Passed Tests")
        ws2.append(["Test ID", "Module", "Test Name", "Priority", "Status", "Execution Time (s)"])
        self._apply_row_style(ws2[1], fill=PatternFill(start_color="166534", end_color="166534", fill_type="solid"), font=HEADER_FONT, alignment=Alignment(horizontal="center"))
        for r in [x for x in self.results if x["status"] == "PASS"]:
            ws2.append([r["test_id"], r["module"], r["name"], r["priority"], r["status"], r.get("duration", 0.0)])
            row_idx = ws2.max_row
            self._apply_row_style(ws2[row_idx])
            ws2.cell(row=row_idx, column=5).fill = PASS_FILL
            ws2.cell(row=row_idx, column=5).font = PASS_FONT
        self._auto_adjust_columns(ws2)

        # Sheet 3: Failed Tests
        ws3 = wb.create_sheet(title="Failed Tests")
        ws3.append(["Test ID", "Module", "Test Name", "Priority", "Failure Reason", "Screenshot", "Execution Time (s)"])
        self._apply_row_style(ws3[1], fill=PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid"), font=HEADER_FONT, alignment=Alignment(horizontal="center"))
        for r in [x for x in self.results if x["status"] == "FAIL"]:
            ws3.append([r["test_id"], r["module"], r["name"], r["priority"], r.get("error_message", "N/A"), r.get("screenshot", "N/A"), r.get("duration", 0.0)])
            row_idx = ws3.max_row
            self._apply_row_style(ws3[row_idx])
        self._auto_adjust_columns(ws3)

        # Sheet 4: Skipped Tests
        ws4 = wb.create_sheet(title="Skipped Tests")
        ws4.append(["Test ID", "Module", "Test Name", "Priority", "Skip Reason"])
        self._apply_row_style(ws4[1], fill=PatternFill(start_color="854D0E", end_color="854D0E", fill_type="solid"), font=HEADER_FONT, alignment=Alignment(horizontal="center"))
        for r in [x for x in self.results if x["status"] in ("SKIPPED", "BLOCKED")]:
            ws4.append([r["test_id"], r["module"], r["name"], r["priority"], r.get("skip_reason", "Precondition unmet")])
            row_idx = ws4.max_row
            self._apply_row_style(ws4[row_idx])
        self._auto_adjust_columns(ws4)

        # Sheet 5: Execution Metrics
        ws5 = wb.create_sheet(title="Execution Metrics")
        ws5.append(["Metric", "Value", "Unit / Context"])
        self._apply_row_style(ws5[1], fill=HEADER_FILL, font=HEADER_FONT, alignment=Alignment(horizontal="center"))
        metrics_data = [
            ("Total Test Cases", self.metrics.get("total", 0), "Total defined test cases"),
            ("Executed Tests", self.metrics.get("executed", 0), "Total tests run"),
            ("Passed Tests", self.metrics.get("passed", 0), "Tests verified successfully"),
            ("Failed Tests", self.metrics.get("failed", 0), "Failed validations"),
            ("Skipped Tests", self.metrics.get("skipped", 0), "Skipped / Blocked"),
            ("Pass Percentage", f"{self.metrics.get('pass_rate', 0.0):.2f}%", "Overall success rate"),
            ("Execution Duration", f"{self.metrics.get('duration_sec', 0.0):.2f}", "Seconds"),
            ("Base URL Tested", self.metrics.get("base_url", ""), "Target live GitHub Pages URL"),
            ("Browser", self.metrics.get("browser", "Chrome Headless"), "Execution Browser Engine"),
            ("Timestamp", self.metrics.get("timestamp", datetime.now().isoformat()), "Execution date & time"),
        ]
        for m in metrics_data:
            ws5.append(list(m))
            self._apply_row_style(ws5[ws5.max_row])
        self._auto_adjust_columns(ws5)

        # Sheet 6: Defect Summary
        ws6 = wb.create_sheet(title="Defect Summary")
        ws6.append(["Defect ID", "Test ID", "Module", "Severity", "Failure Reason", "Stack Trace Snippet"])
        self._apply_row_style(ws6[1], fill=HEADER_FILL, font=HEADER_FONT, alignment=Alignment(horizontal="center"))
        defect_count = 1
        for r in [x for x in self.results if x["status"] == "FAIL"]:
            stack_snippet = (r.get("stack_trace") or "")[:200]
            ws6.append([f"DEF-{defect_count:03d}", r["test_id"], r["module"], r.get("priority", "High"), r.get("error_message", "Assertion failed"), stack_snippet])
            self._apply_row_style(ws6[ws6.max_row])
            defect_count += 1
        if defect_count == 1:
            ws6.append(["None", "N/A", "N/A", "N/A", "Zero defects detected during execution", ""])
            self._apply_row_style(ws6[ws6.max_row])
        self._auto_adjust_columns(ws6)

        # Save to both Test Results/Excel and automation/reports
        out_file_1 = EXCEL_RESULTS_DIR / "Automation_Test_Report.xlsx"
        out_file_2 = REPORTS_DIR / "Automation_Test_Report.xlsx"
        wb.save(str(out_file_1))
        wb.save(str(out_file_2))
        logger.info(f"Generated main Excel report: {out_file_1}")
        return out_file_1

    def generate_filtered_report(self, filter_name: str, filtered_items: list[dict], filename: str) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{filter_name} Test Cases"
        headers = ["Test ID", "Module", "Test Name", "Priority", "Status", "Details", "Execution Time (s)"]
        ws.append(headers)
        self._apply_row_style(ws[1], fill=HEADER_FILL, font=HEADER_FONT, alignment=Alignment(horizontal="center"))

        for r in filtered_items:
            details = r.get("error_message") or r.get("skip_reason") or "Verified successfully"
            ws.append([r["test_id"], r["module"], r["name"], r["priority"], r["status"], details, r.get("duration", 0.0)])
            row_idx = ws.max_row
            self._apply_row_style(ws[row_idx])
            status_cell = ws.cell(row=row_idx, column=5)
            if r["status"] == "PASS":
                status_cell.fill = PASS_FILL
                status_cell.font = PASS_FONT
            elif r["status"] == "FAIL":
                status_cell.fill = FAIL_FILL
                status_cell.font = FAIL_FONT
            else:
                status_cell.fill = SKIP_FILL
                status_cell.font = SKIP_FONT

        if not filtered_items:
            ws.append(["N/A", "N/A", f"No {filter_name.lower()} test cases", "N/A", "N/A", "", 0.0])
            self._apply_row_style(ws[ws.max_row])

        self._auto_adjust_columns(ws)
        out_file_1 = EXCEL_RESULTS_DIR / filename
        out_file_2 = REPORTS_DIR / filename
        wb.save(str(out_file_1))
        wb.save(str(out_file_2))
        logger.info(f"Generated {filter_name} Excel report: {out_file_1}")
        return out_file_1

    def generate_summary_report(self, filename: str) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        
        ws.append(["AgroSentry-AI Live GitHub Pages E2E Execution Summary", ""])
        ws.merge_cells("A1:B1")
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="1E293B")
        ws.append([])

        summary_rows = [
            ("Execution Status", "PASSED" if self.metrics.get("pass_rate", 0) >= 95.0 else "FAILED"),
            ("Target Deployment URL", self.metrics.get("base_url", "")),
            ("Execution Date", self.metrics.get("timestamp", datetime.now().isoformat())),
            ("Total Test Cases", self.metrics.get("total", 0)),
            ("Executed", self.metrics.get("executed", 0)),
            ("Passed", self.metrics.get("passed", 0)),
            ("Failed", self.metrics.get("failed", 0)),
            ("Skipped", self.metrics.get("skipped", 0)),
            ("Pass Rate", f"{self.metrics.get('pass_rate', 0.0):.2f}%"),
            ("Execution Time (s)", f"{self.metrics.get('duration_sec', 0.0):.2f}"),
        ]

        for k, v in summary_rows:
            ws.append([k, v])
            row_idx = ws.max_row
            ws.cell(row=row_idx, column=1).font = Font(name="Calibri", bold=True)
            self._apply_row_style(ws[row_idx])

        self._auto_adjust_columns(ws)
        out_file_1 = EXCEL_RESULTS_DIR / filename
        out_file_2 = REPORTS_DIR / filename
        wb.save(str(out_file_1))
        wb.save(str(out_file_2))
        logger.info(f"Generated summary Excel report: {out_file_1}")
        return out_file_1
